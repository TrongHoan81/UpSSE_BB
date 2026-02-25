import io
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from .utils import (
    normalize_mst, format_tax_code, clean_string, to_float, 
    to_tax_rate_float, calculate_similarity
)

def load_bvmt_mapping():
    """Tải đơn giá thuế Bảo vệ môi trường từ Data/BVMT.xlsx."""
    mapping = {}
    path = os.path.join('Data', 'BVMT.xlsx')
    if not os.path.exists(path): return mapping
    try:
        df = pd.read_excel(path, header=None)
        for i, row in df.iterrows():
            name = clean_string(row.iloc[0]).upper()
            if name == 'TÊN MẶT HÀNG' or not name: continue
            mapping[name] = to_float(row.iloc[1])
    except: pass
    return mapping

def load_makho_mapping():
    """Tải mã kho từ Data/MaKho_MaVV.xlsx."""
    mapping = {}
    path = os.path.join('Data', 'MaKho_MaVV.xlsx')
    if not os.path.exists(path): return mapping
    try:
        df = pd.read_excel(path, header=None)
        for i, row in df.iterrows():
            ten_kho = clean_string(row.iloc[0]).upper()
            ma_kho = str(row.iloc[1]).strip()
            if ten_kho and ten_kho != 'TÊN KHO XUẤT HÀNG':
                mapping[ten_kho] = ma_kho
    except: pass
    return mapping

def load_vuviec_matrix():
    """Tải ma trận mã vụ việc Kho x Mặt hàng từ Data/MaKho_MaVV.xlsx."""
    matrix = {}
    path = os.path.join('Data', 'MaKho_MaVV.xlsx')
    if not os.path.exists(path): return matrix
    try:
        df = pd.read_excel(path, header=None)
        if len(df) < 2: return matrix
        product_headers = {}
        for col_idx in range(2, len(df.columns)):
            prod_name = clean_string(df.iloc[1, col_idx]).upper()
            if prod_name: product_headers[col_idx] = prod_name
        for row_idx in range(2, len(df)):
            ten_kho = clean_string(df.iloc[row_idx, 0]).upper()
            if not ten_kho: continue
            matrix[ten_kho] = {}
            for col_idx, prod_name in product_headers.items():
                vv_code = str(df.iloc[row_idx, col_idx]).strip()
                if vv_code and vv_code.lower() != 'nan':
                    matrix[ten_kho][prod_name] = vv_code
    except: pass
    return matrix

def load_mahh_mapping():
    """Tải danh mục mặt hàng và hệ số VCF từ Data/MaHH.xlsx."""
    code_mapping = {}
    vcf_mapping = {}
    path = os.path.join('Data', 'MaHH.xlsx')
    if not os.path.exists(path): return code_mapping, vcf_mapping
    try:
        df = pd.read_excel(path, header=None)
        for i, row in df.iterrows():
            ten_hang = clean_string(row.iloc[0]).upper()
            if not ten_hang or ten_hang == 'TÊN MẶT HÀNG': continue
            code_mapping[ten_hang] = str(row.iloc[1]).strip()
            vcf_mapping[ten_hang] = {
                'winter': to_float(row.iloc[2]),
                'summer': to_float(row.iloc[3])
            }
    except: pass
    return code_mapping, vcf_mapping

def load_accounts_mapping():
    """Tải danh sách tài khoản định khoản từ Data/DanhSachTaiKhoan.xlsx."""
    path = os.path.join('Data', 'DanhSachTaiKhoan.xlsx')
    acc = {
        'tk_no': "", 'tk_thue_co': "", 'tk_thue_no': "", 
        'tk_dt': "", 'tk_kho': "", 'tk_gia_von': "",
        'tk_dt_bvmt': ""
    }
    if not os.path.exists(path): return acc
    try:
        df = pd.read_excel(path, header=None)
        acc['tk_no'] = str(df.iloc[1, 1]).strip() if len(df) > 1 else ""
        acc['tk_thue_co'] = str(df.iloc[2, 1]).strip() if len(df) > 2 else ""
        acc['tk_thue_no'] = str(df.iloc[3, 1]).strip() if len(df) > 3 else ""
        acc['tk_dt'] = str(df.iloc[4, 1]).strip() if len(df) > 4 else ""
        acc['tk_kho'] = str(df.iloc[5, 1]).strip() if len(df) > 5 else ""
        acc['tk_gia_von'] = str(df.iloc[6, 1]).strip() if len(df) > 6 else ""
        acc['tk_dt_bvmt'] = str(df.iloc[9, 1]).strip() if len(df) > 9 else ""
    except: pass
    return acc

def merge_and_fill_template(invoice_data, bm19_data, template_path, manual_date=None):
    """
    Ghép nối và xử lý định khoản. 
    Bổ sung điền giá trị "1" vào cột AD cho tất cả các dòng.
    """
    print(f"\n[LOG] --- BẮT ĐẦU XỬ LÝ GHÉP NỐI ---")
    clean_manual_date = str(manual_date).strip() if manual_date else None

    bvmt_map = load_bvmt_mapping()
    makho_map = load_makho_mapping()
    vuviec_matrix = load_vuviec_matrix()
    mahh_code_map, mahh_vcf_map = load_mahh_mapping()
    acc_map = load_accounts_mapping()
    
    customer_map = {}
    dskh_path = os.path.join('Data', 'DSKH.xlsx')
    if os.path.exists(dskh_path):
        try:
            df_kh = pd.read_excel(dskh_path)
            for _, r in df_kh.iterrows():
                mst = normalize_mst(r.iloc[1]); ma_sse = str(r.iloc[2]).strip()
                if mst: customer_map[mst] = ma_sse
        except: pass

    if os.path.exists(template_path): wb = load_workbook(template_path); ws = wb.active
    else: wb = Workbook(); ws = wb.active

    # 1. CHUYỂN ĐỔI NGÀY SANG DATE OBJECT
    final_date_obj = None
    if clean_manual_date:
        try:
            final_date_obj = datetime.strptime(clean_manual_date, '%d/%m/%Y')
            print(f"[LOG] Chuyển đổi sang Datetime thành công: {final_date_obj}")
        except Exception as e:
            print(f"[LOG] CẢNH BÁO: Không thể parse '{clean_manual_date}'. Lỗi: {e}")
            final_date_obj = clean_manual_date

    # --- GIAI ĐOẠN 1: DÒNG HÀNG HÓA ---
    curr_row = 6
    for item in invoice_data:
        mat_hang_hd = item['mat_hang'].upper().strip()
        ten_kh_hd = item['ten_kh']
        sl_hd = round(item['so_luong'], 3)
        ten_kho_hd = item['kho_xuat_bkhd'].upper().strip()

        # Tra cứu BM19
        nhiet_do = None; ty_trong = None
        matches = [b for b in bm19_data if b['mat_hang'] == mat_hang_hd and b['so_luong'] == sl_hd]
        best_match = None; highest_score = 0.0
        for p in matches:
            score = calculate_similarity(ten_kh_hd, p['ten_kh'])
            if score > highest_score: highest_score = score; best_match = p
        if best_match and highest_score >= 0.9:
            nhiet_do = best_match['temp']; ty_trong = best_match['dens']

        don_gia_bvmt = bvmt_map.get(mat_hang_hd, 0)
        vat_rate = to_tax_rate_float(item['vat_raw'])
        tien_thue_bvmt = sl_hd * don_gia_bvmt
        thue_tren_bvmt = tien_thue_bvmt * vat_rate
        
        tien_thue_line1 = item['tien_thue_total_bkhd'] - thue_tren_bvmt
        tien_hang_line1 = item['thanh_tien_total_bkhd'] - tien_thue_bvmt
        gia_ban_line1 = item['don_gia_bkhd'] - don_gia_bvmt

        # Ghi dữ liệu
        ws.cell(row=curr_row, column=1).value = customer_map.get(item['mst_key'], "")
        ws.cell(row=curr_row, column=2).value = ten_kh_hd
        ws.cell(row=curr_row, column=3).value = final_date_obj
        ws.cell(row=curr_row, column=4).value = f"{item['ky_hieu'][:5]}{item['so_hd'][-5:].zfill(5)}"
        ws.cell(row=curr_row, column=5).value = f"{item['mau_so']}{item['ky_hieu']}"
        ws.cell(row=curr_row, column=6).value = f"Xuất bán {item['mat_hang']} theo HĐ {item['so_hd']} số lượng {sl_hd} lít"
        ws.cell(row=curr_row, column=7).value = acc_map['tk_no']
        ws.cell(row=curr_row, column=11).value = ty_trong 
        ws.cell(row=curr_row, column=12).value = nhiet_do 
        ws.cell(row=curr_row, column=14).value = format_tax_code(item['vat_raw'])
        ws.cell(row=curr_row, column=15).value = tien_thue_line1
        ws.cell(row=curr_row, column=16).value = acc_map['tk_thue_co']
        ws.cell(row=curr_row, column=17).value = acc_map['tk_thue_no']
        ws.cell(row=curr_row, column=18).value = mahh_code_map.get(mat_hang_hd, "")
        ws.cell(row=curr_row, column=19).value = item['dvt']
        ws.cell(row=curr_row, column=20).value = makho_map.get(ten_kho_hd, "")
        ws.cell(row=curr_row, column=21).value = sl_hd
        ws.cell(row=curr_row, column=22).value = gia_ban_line1
        ws.cell(row=curr_row, column=23).value = tien_hang_line1
        ws.cell(row=curr_row, column=24).value = acc_map['tk_dt']
        ws.cell(row=curr_row, column=25).value = acc_map['tk_kho']
        ws.cell(row=curr_row, column=26).value = acc_map['tk_gia_von']
        
        ma_vv = ""
        if ten_kho_hd in vuviec_matrix: ma_vv = vuviec_matrix[ten_kho_hd].get(mat_hang_hd, "")
        ws.cell(row=curr_row, column=28).value = ma_vv
        
        # ĐIỀN CỘT AD (CỘT 30) = 1
        ws.cell(row=curr_row, column=30).value = 1
        
        curr_row += 1

    # --- GIAI ĐOẠN 2: DÒNG THUẾ BVMT ---
    for item in invoice_data:
        mat_hang_hd = item['mat_hang'].upper().strip()
        ten_kho_hd = item['kho_xuat_bkhd'].upper().strip()
        sl_hd = round(item['so_luong'], 3)
        don_gia_bvmt = bvmt_map.get(mat_hang_hd, 0)
        vat_rate = to_tax_rate_float(item['vat_raw'])
        tien_thue_bvmt = sl_hd * don_gia_bvmt
        thue_tren_bvmt = tien_thue_bvmt * vat_rate
        ma_kho = makho_map.get(ten_kho_hd, "")
        ma_vv = ""
        if ten_kho_hd in vuviec_matrix: ma_vv = vuviec_matrix[ten_kho_hd].get(mat_hang_hd, "")

        ws.cell(row=curr_row, column=1).value = customer_map.get(item['mst_key'], "")
        ws.cell(row=curr_row, column=2).value = item['ten_kh']
        ws.cell(row=curr_row, column=3).value = final_date_obj
        ws.cell(row=curr_row, column=4).value = f"{item['ky_hieu'][:5]}{item['so_hd'][-5:].zfill(5)}"
        ws.cell(row=curr_row, column=5).value = f"{item['mau_so']}{item['ky_hieu']}"
        ws.cell(row=curr_row, column=6).value = "" 
        ws.cell(row=curr_row, column=7).value = acc_map['tk_no'] 
        ws.cell(row=curr_row, column=14).value = format_tax_code(item['vat_raw']) 
        ws.cell(row=curr_row, column=15).value = thue_tren_bvmt 
        ws.cell(row=curr_row, column=16).value = acc_map['tk_thue_co']
        ws.cell(row=curr_row, column=17).value = acc_map['tk_thue_no']
        ws.cell(row=curr_row, column=18).value = "TMT" 
        ws.cell(row=curr_row, column=19).value = item['dvt']
        ws.cell(row=curr_row, column=20).value = ma_kho
        ws.cell(row=curr_row, column=21).value = sl_hd
        ws.cell(row=curr_row, column=22).value = don_gia_bvmt 
        ws.cell(row=curr_row, column=23).value = tien_thue_bvmt 
        ws.cell(row=curr_row, column=24).value = acc_map['tk_dt_bvmt'] 
        ws.cell(row=curr_row, column=25).value = acc_map['tk_kho']
        ws.cell(row=curr_row, column=26).value = acc_map['tk_gia_von']
        ws.cell(row=curr_row, column=28).value = ma_vv
        
        # ĐIỀN CỘT AD (CỘT 30) = 1
        ws.cell(row=curr_row, column=30).value = 1
        
        curr_row += 1

    # --- GIAI ĐOẠN 3: QUÉT ĐỊNH DẠNG HẬU KỲ ---
    print(f"[LOG] Bắt đầu quét hậu kỳ định dạng cho {ws.max_row - 5} dòng...")
    for r_idx in range(6, ws.max_row + 1):
        cell_date = ws.cell(row=r_idx, column=3)
        val = cell_date.value
        if isinstance(val, str):
            clean_val = val.strip()
            try:
                val = datetime.strptime(clean_val, '%d/%m/%Y')
                cell_date.value = val
            except: pass
        if isinstance(val, datetime):
            cell_date.number_format = 'dd/mm/yyyy'
        ws.cell(row=r_idx, column=18).number_format = '@'

    print(f"[LOG] --- XỬ LÝ HOÀN TẤT ---")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output