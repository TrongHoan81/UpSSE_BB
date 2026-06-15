import io
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from .utils import clean_string, to_float

def load_makho_dieuchuyen():
    """Tải mã kho nhập từ Data/MaKho.xlsx."""
    mapping = {}
    path = os.path.join('Data', 'MaKho.xlsx')
    if not os.path.exists(path): return mapping
    try:
        df = pd.read_excel(path, header=None)
        for i, row in df.iterrows():
            ten_kho = clean_string(row.iloc[0]).upper()
            ma_kho = str(row.iloc[1]).strip()
            if ten_kho:
                mapping[ten_kho] = ma_kho
    except: pass
    return mapping

def load_mahh_dieuchuyen():
    """Tải mã hàng hóa từ Data/MaHH.xlsx."""
    mapping = {}
    path = os.path.join('Data', 'MaHH.xlsx')
    if not os.path.exists(path): return mapping
    try:
        df = pd.read_excel(path, header=None)
        for i, row in df.iterrows():
            ten_hang = clean_string(row.iloc[0]).upper()
            ma_hang = str(row.iloc[1]).strip()
            if ten_hang and ten_hang != 'TÊN MẶT HÀNG': 
                mapping[ten_hang] = ma_hang
    except: pass
    return mapping

def load_kyhieu_dieuchuyen():
    """Lấy ký hiệu từ ô A1 của Data/KyHieu.xlsx."""
    path = os.path.join('Data', 'KyHieu.xlsx')
    if not os.path.exists(path): return ""
    try:
        df = pd.read_excel(path, header=None)
        val = str(df.iloc[0, 0]).strip()
        return val if val.lower() != 'nan' else ""
    except: return ""

def load_vuviec_dieuchuyen():
    """
    Tải mã vụ việc từ Data/MaKho_MaVV.xlsx 
    (Tra cứu Tên mặt hàng ở dòng 2 -> Lấy Vụ việc ở dòng 3)
    """
    mapping = {}
    path = os.path.join('Data', 'MaKho_MaVV.xlsx')
    if not os.path.exists(path): return mapping
    try:
        df = pd.read_excel(path, header=None)
        # Kiểm tra xem file có ít nhất 3 dòng không (Dòng 1: index 0, Dòng 2: index 1, Dòng 3: index 2)
        if len(df) >= 3:
            for col_idx in range(len(df.columns)):
                prod_name = clean_string(df.iloc[1, col_idx]).upper()
                vv_code = str(df.iloc[2, col_idx]).strip()
                if prod_name and vv_code and vv_code.lower() != 'nan':
                    mapping[prod_name] = vv_code
    except: pass
    return mapping

def process_transfer_data(file_storage, template_path):
    """
    Trích xuất dữ liệu từ BM19 và tạo Phiếu xuất điều chuyển.
    Đã bổ sung đầy đủ logic điền từ cột A đến AI.
    """
    print(f"\n[LOG] --- BẮT ĐẦU XỬ LÝ PHIẾU XUẤT ĐIỀU CHUYỂN ---")
    
    makho_map = load_makho_dieuchuyen()
    mahh_map = load_mahh_dieuchuyen()
    ky_hieu_val = load_kyhieu_dieuchuyen()
    vuviec_map = load_vuviec_dieuchuyen()
    
    if os.path.exists(template_path): 
        wb = load_workbook(template_path)
        ws = wb.active
    else: 
        wb = Workbook()
        ws = wb.active
        
    try:
        df = pd.read_excel(file_storage, header=None)
        
        # Tìm dòng tiêu đề
        header_idx = -1
        for i, row in df.iterrows():
            vals = [str(c).strip().upper() for c in row if pd.notna(c)]
            if 'MẶT HÀNG' in vals and 'KHÁCH HÀNG' in vals:
                header_idx = i
                break
                
        if header_idx == -1:
            raise ValueError("Không tìm thấy dòng tiêu đề chuẩn trong file BM19.")

        curr_row = 6
        for idx in range(header_idx + 1, len(df)):
            row = df.iloc[idx]
            
            khach_hang_raw = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ""
            
            # 1. ĐIỀU KIỆN LỌC
            if not khach_hang_raw.upper().startswith("CHXD"):
                continue
                
            mat_hang_raw = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ""
            if not mat_hang_raw:
                continue

            ngay_xuat = row.iloc[2] if pd.notna(row.iloc[2]) else ""
            so_tc = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ""
            nhiet_do = row.iloc[7] if pd.notna(row.iloc[7]) else ""
            ty_trong = row.iloc[8] if pd.notna(row.iloc[8]) else ""
            so_luong = row.iloc[12] if pd.notna(row.iloc[12]) else 0
            
            so_luong_format = f"{to_float(so_luong):g}" if so_luong else "0"
            dien_giai = f"Xuất điều chuyển {mat_hang_raw} {so_luong_format}"

            khach_hang_upper = clean_string(khach_hang_raw).upper()
            mat_hang_upper = clean_string(mat_hang_raw).upper()

            # 2. GHI DỮ LIỆU VÀO TEMPLATE (Cột A - S)
            ws.cell(row=curr_row, column=1).value = "KHOTC01"                      # A
            ws.cell(row=curr_row, column=2).value = "Kho trung chuyển"             # B
            ws.cell(row=curr_row, column=3).value = makho_map.get(khach_hang_upper, "") # C
            ws.cell(row=curr_row, column=4).value = khach_hang_raw                 # D
            ws.cell(row=curr_row, column=5).value = ""                             # E
            ws.cell(row=curr_row, column=6).value = 3                              # F
            ws.cell(row=curr_row, column=7).value = ngay_xuat                      # G
            ws.cell(row=curr_row, column=8).value = so_tc                          # H
            ws.cell(row=curr_row, column=9).value = dien_giai                      # I
            ws.cell(row=curr_row, column=10).value = 2                             # J
            ws.cell(row=curr_row, column=11).value = mahh_map.get(mat_hang_upper, "") # K
            ws.cell(row=curr_row, column=12).value = mat_hang_raw                  # L
            ws.cell(row=curr_row, column=13).value = "lít"                         # M
            ws.cell(row=curr_row, column=14).value = ""                            # N
            ws.cell(row=curr_row, column=15).value = ""                            # O
            ws.cell(row=curr_row, column=16).value = ""                            # P
            ws.cell(row=curr_row, column=17).value = so_luong                      # Q
            ws.cell(row=curr_row, column=18).value = ty_trong                      # R
            ws.cell(row=curr_row, column=19).value = nhiet_do                      # S

            # 3. GHI DỮ LIỆU BỔ SUNG (Cột T - AI)
            ws.cell(row=curr_row, column=20).value = ""                            # T: Giá đích danh
            ws.cell(row=curr_row, column=21).value = ""                            # U: Giá
            ws.cell(row=curr_row, column=22).value = ""                            # V: Tiền
            ws.cell(row=curr_row, column=23).value = ""                            # W: Mã nt
            ws.cell(row=curr_row, column=24).value = ""                            # X: Tỷ giá
            ws.cell(row=curr_row, column=25).value = "1561"                        # Y: Mã nx
            ws.cell(row=curr_row, column=26).value = "1561"                        # Z: Tk nợ
            ws.cell(row=curr_row, column=27).value = "1561"                        # AA: Tk có
            ws.cell(row=curr_row, column=28).value = vuviec_map.get(mat_hang_upper, "") # AB: Vụ việc (Theo dòng 2 & 3 file MaKho_MaVV)
            ws.cell(row=curr_row, column=29).value = ""                            # AC: Bộ phận
            ws.cell(row=curr_row, column=30).value = ""                            # AD: Lsx
            ws.cell(row=curr_row, column=31).value = ""                            # AE: Sản phẩm
            ws.cell(row=curr_row, column=32).value = ""                            # AF: Hợp đồng
            ws.cell(row=curr_row, column=33).value = ""                            # AG: Phí
            ws.cell(row=curr_row, column=34).value = ""                            # AH: Khế ước
            ws.cell(row=curr_row, column=35).value = ky_hieu_val                   # AI: Ký hiệu (Từ ô A1 file KyHieu.xlsx)
            
            curr_row += 1

        print(f"[LOG] --- XỬ LÝ ĐIỀU CHUYỂN HOÀN TẤT ({curr_row - 6} dòng) ---")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        raise ValueError(f"Lỗi khi xử lý Phiếu điều chuyển: {str(e)}")